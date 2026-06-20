from django.shortcuts import render
from openpyxl import load_workbook
from pathlib import Path
from collections import Counter
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.contrib.auth.models import User

def crear_admin(request):

    if not User.objects.filter(username="Robinsson").exists():

        User.objects.create_superuser(
            username="Robinsson",
            password="Caja-852456"
        )

        return HttpResponse("Administrador creado")

    return HttpResponse("Administrador ya existe")

def usuarios(request):

    cantidad = User.objects.count()

    return HttpResponse(
        f"Cantidad de usuarios: {cantidad}"
    )

def moneda(valor):

    if valor is None:
        return "$ 0,00"

    try:

        valor = float(valor)

        texto = f"{valor:,.2f}"

        texto = texto.replace(",", "X")
        texto = texto.replace(".", ",")
        texto = texto.replace("X", ".")

        return f"$ {texto}"

    except:

        return "$ 0,00"


def moneda_usd(valor):

    if valor is None:
        return "USD 0,00"

    try:

        valor = float(valor)

        texto = f"{valor:,.2f}"

        texto = texto.replace(",", "X")
        texto = texto.replace(".", ",")
        texto = texto.replace("X", ".")

        return f"USD {texto}"

    except:

        return "USD 0,00"

@login_required
def dashboard(request):

    archivo = Path("data/Caja.xlsx")

    wb = load_workbook(
        archivo,
        data_only=True
    )

    hoja = wb["Caja"]

    contexto = {

        "operador": hoja["A3"].value,
        "fecha": hoja["B3"].value,

        "total_dia": moneda(hoja["B8"].value),
        "mercado_libre": moneda(hoja["B9"].value),
        "distribucion": moneda(hoja["B10"].value),
        "total_local": moneda(hoja["B11"].value),
        "local_efectivo": moneda(hoja["B12"].value),
        "saldo_final_uyu": moneda(hoja["B13"].value),

        "saldo_inicial_uyu": moneda(hoja["B4"].value),
        "movimientos_uyu": moneda(hoja["B5"].value),

        "sobrante_uyu": moneda(hoja["B14"].value),
        "faltante_uyu": moneda(hoja["B15"].value),

        "saldo_inicial_usd": moneda_usd(hoja["B6"].value),
        "movimientos_usd": moneda_usd(hoja["B7"].value),

        "saldo_final_usd": moneda_usd(hoja["B16"].value),
        "sobrante_usd": moneda_usd(hoja["B17"].value),
        "faltante_usd": moneda_usd(hoja["B18"].value),

        "observacion": hoja["B19"].value,
    }

    return render(
        request,
        "dashboard.html",
        contexto
    )

@login_required
def asistencias(request):

    archivo = Path("data/Caja.xlsx")

    wb = load_workbook(
        archivo,
        data_only=True
    )

    hoja = wb["Ast. Caja"]

    registros = []

    fila = 2

    responsables = []

    while hoja[f"A{fila}"].value:

        responsable = hoja[f"B{fila}"].value or "Sin Operador"

        registros.append({

            "fecha": hoja[f"A{fila}"].value,
            "responsable": responsable,
            "dia": hoja[f"C{fila}"].value,

        })

        responsables.append(responsable)

        fila += 1

    contador = Counter(responsables)

    operador_principal = ""

    dias_operador_principal = 0

    if contador:

        operador_principal = contador.most_common(1)[0][0]

        dias_operador_principal = contador.most_common(1)[0][1]

    ranking_operadores = []

    max_dias = max(contador.values()) if contador else 1

    for nombre, dias in contador.most_common():

        porcentaje = int((dias / max_dias) * 100)

        ranking_operadores.append({

            "nombre": nombre,
            "dias": dias,
            "porcentaje": porcentaje

        })

    contexto = {

        "registros": registros,

        "dias_operados": len(registros),

        "responsables_activos": len(contador),

        "operador_principal": operador_principal,

        "dias_operador_principal": dias_operador_principal,

        "ranking_operadores": ranking_operadores,

    }

    return render(
        request,
        "asistencias.html",
        contexto
    )

